import os
import re
import openpyxl
import requests
import pandas as pd
from os import getcwd, path
from parsel import Selector

def clean(text):
    '''remove extra spaces & junk character'''
    text = re.sub(r'\n+','',text)
    text = re.sub(r'\s+',' ',text)
    text = re.sub(r'\r+','',text)
    return text.strip()

file_path = getcwd()
file_name = input("Enter file name : ") + ".xlsx"
output_filename = f"{file_path}\\{file_name}"
wb = openpyxl.load_workbook(output_filename)
data_sheet = wb.active
data_sheet["D1"] = "headers"
data_sheet["E1"] = "paragraph"
data_sheet["F1"] = "paragraph_content"
for i in range(1, data_sheet.max_row + 1):
    item = {}
    url = data_sheet.cell(row=i, column=1).value
    payload = {}
    headers = {
    'authority': 'www.lenovo.com',
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'accept-language': 'en-US,en;q=0.9',
    'cache-control': 'max-age=0',
    'sec-ch-ua': '"Chromium";v="112", "Google Chrome";v="112", "Not:A-Brand";v="99"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-fetch-dest': 'document',
    'sec-fetch-mode': 'navigate',
    'sec-fetch-site': 'none',
    'sec-fetch-user': '?1',
    'upgrade-insecure-requests': '1',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36'
    }

    response = requests.request("GET", url, headers=headers, data=payload)
    response_text = Selector(text=response.text)
    full_content = response_text.xpath('//div[@class="lenovo_body"]/main[@class="main_content"]')
    item['headers'] =[clean(i) for i in full_content.xpath('//h1//text()|//h2//text()|//h3//text()|//h4//text()').getall() if i]
    item['paragraph'] =[clean(i)  for i in full_content.xpath('//ul//li//text()|//p//text()|//span//text()|//div[@class="copy-block"]//text()').getall() if i]
    item['url'] = url
    df = pd.DataFrame([item])
    if not os.path.isfile("link_collection.csv"):
        df.to_csv("link_collection.csv",index=False,mode="a",header=True,encoding="utf_8_sig",)
    else:  # else it exists so append without writing the header
        df.to_csv("link_collection.csv",index=False,mode="a",header=False,encoding="utf_8_sig",)
df = pd.read_csv('link_collection.csv', sep=',')
df.to_excel('link_collection.xlsx', index=False)
        
        



        

