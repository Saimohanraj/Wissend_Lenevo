import openpyxl, traceback
import os, os.path, time, json, pathlib
import requests, lxml, re, math
from bs4 import BeautifulSoup
from random import randrange
from lxml.html import fromstring
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from parsel import Selector

########################################################### [ UNDECTED CHORMEDRIVER ] ###########################################################
current_directory = os.getcwd()+"\\"
options = webdriver.ChromeOptions()
options.add_argument("start-maximized")
options.add_argument('disable-extensions')
options.add_argument('disable-gpu')
options.add_argument('disable-infobars')
options.add_argument('disable-ntp-most-likely-favicons-from-server')
options.add_argument('disable-login-animations')
options.add_argument('disable-popup-blocking')
options.add_argument('disable-images')
options.add_argument('log-level=0')
options.add_argument('log-level=1')
options.add_argument('log-level=2')
options.add_argument('log-level=3')
options.add_experimental_option("prefs",{'profile.managed_default_content_settings.images': 2})
techDriver = webdriver.Chrome(options=options)
time.sleep(2)
########################################################### [ UNDECTED CHORMEDRIVER ] ###########################################################

##################### General Content ############################
now_time = time.localtime()
now_date = time.strftime("%d-%m-%Y", now_time)
time_on = time.strftime("%I:%M %p", now_time)
timer_start = time.time()
print('\n'+'Start time: '+time_on)

website = 'Lenovo'
connect_domain = 'https://www.lenovo.com'
selective_category = ['laptop', 'desktop', 'workstation', 'tablet']
lenovo_headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36 Edg/113.0.1774.57'}

def url_correction(url):
	global connect_domain
	if url.startswith('/') == True:
		return 'https:'+url if 'lenovo.com' in url.lower() else connect_domain+url
	else:
		if 'http' in url:
			return url
		else:
			return connect_domain+'/'+url if url[0].isalpha() == True else url
##################### General Content ############################

#################### workbook module ############################
##################### input
get_file = input('Enter the ["Lenovo Countries"] input file name without extensions : ')
path = current_directory+'/'+get_file+'.xlsx'
con_in_wb = openpyxl.load_workbook(path)
con_in_ws = con_in_wb.active
con_max_rows = con_in_ws.max_row+1

lenovo_contries = []
[lenovo_contries.append((con_in_ws.cell(i, 1).value, con_in_ws.cell(i, 2).value)) for i in range(2, con_max_rows)]

##################### output
def sheet_headers(ws):
	ws['A1'].value = 'Category'
	ws['B1'].value = 'Country'
	ws['C1'].value = 'Language'
	ws['D1'].value = 'Brand'
	ws['E1'].value = 'Part Title'
	ws['F1'].value = 'Part Number'
	ws['G1'].value = 'Classification Attributes'
	ws['H1'].value = 'String Value'
	ws['I1'].value = 'Link'
################### workbook module ############################

############################################################### [ Common - Land Page Execution ] ###############################################################
def multi_product(soup, writing_dict, Land_url):
	common_title_for_multi = soup.select('#tab-customize h2.tabbedBrowse-title')[0].text.strip() if len(soup.select('#tab-customize h2.tabbedBrowse-title')) != 0 else ''
	hmc_products_list = soup.select('ol.tabbedBrowse-productListings li[data-code]')
	flash_products_list = soup.select('ul.skeleton_product li.product_item')
	if len(hmc_products_list) != 0:
		for hmc_product in hmc_products_list:
			title = hmc_product.select('h3.tabbedBrowse-productListing-title')[0].text.strip() if len(hmc_product.select('h3.tabbedBrowse-productListing-title')) != 0 else common_title_for_multi
			if 'Part Number' in title:
				title = title.split('Part Number')[0].strip()
			else:
				title = title.split(':')[0].strip()
			if '\n' in title:
				title = title.split('\n')[0]
			
			partnumber = hmc_product['data-code']
			brand_name = soup.select('[name="brand"]')[0]['content']
			if partnumber != '' and title != '':
				writing_dict[partnumber+'|'+title+'|'+brand_name+'|'+Land_url] = []
				if len(hmc_product.select('.expandableContent dl')) != 0:
					for multiple_content in hmc_product.select('.expandableContent dl'):
						try:
							content = re.findall(r'data-term=\"(.*?)\".*?<dd.*?>(.*?)<\/dd>', str(str(multiple_content).replace('\n','')))
						except:
							content = re.findall(r'<dt.*?>(.*?)<\/dt><dd.*?>(.*?)<\/dd>', str(str(multiple_content).replace('\n','')))
						for ck, ci in enumerate(content):
							if '<img' in list(ci)[1]:
								content[ck] = (list(ci)[0], '')
						if len(content) != 0:
							writing_dict[partnumber+'|'+title+'|'+brand_name+'|'+Land_url] = writing_dict[partnumber+'|'+title+'|'+brand_name+'|'+Land_url] + content
	elif len(flash_products_list) != 0:
		for flash_product in flash_products_list:
			title = flash_product.select('.product_title span[id^="title"]')[0].text.strip() if len(flash_product.select('.product_title span[id^="title"]')) != 0 else common_title_for_multi
			if 'Part Number' in title:
				title = title.split('Part Number')[0].strip()
			else:
				title = title.split(':')[0].strip()
			if '\n' in title:
				title = title.split('\n')[0]

			partnumber = ''
			partnumber_data = flash_product['data-adobe-params']
			if partnumber_data != '':
				partnumber = re.findall(r'\"productNubmer\":\"(.*?)\",\"',str(partnumber_data))[0]
			else:
				partnumber = flash_product.select('.part_number span:nth-child(2)')[0].text.strip() if len(flash_product.select('.part_number span:nth-child(2)')) != 0 else ''
			brand_name = soup.select('[name="brand"]')[0]['content'] if len(soup.select('[name="brand"]')) != 0 else ''
			
			if partnumber != '' and title != '':
				writing_dict[partnumber+'|'+title+'|'+brand_name+'|'+Land_url] = []
				if len(flash_product.select('.key_details ul')) != 0:
					for multiple_content in flash_product.select('.key_details ul li'):
						if len(multiple_content.select('span')) == 2:
							flash_attribute_name = multiple_content.select('span')[0].text.strip().replace(' :','')
							flash_attribute_val = multiple_content.select('span')[1].text.strip()
							writing_dict[partnumber+'|'+title+'|'+brand_name+'|'+Land_url].append((flash_attribute_name, flash_attribute_val))
						else:	
							writing_dict[partnumber+'|'+title+'|'+brand_name+'|'+Land_url].append(('', ''))

def single_product(soup, writing_dict, Land_url):
	title = partnumber = ''
	
	platform = soup.select('meta[name="platform"][content="Flash"]')
	if len(platform) != 0:
		title = soup.select('.banner_content_desc h2')[0].text.strip() if len(soup.select('.banner_content_desc h2')) != 0 else ''
		partnumber = soup.select('.banner_content_desc [data-product-code]')[0]['data-product-code'] if len(soup.select('.banner_content_desc [data-product-code]')) != 0 else ''
		brand_name = soup.select('[name="brand"]')[0]['content'] if len(soup.select('[name="brand"]')) != 0 else ''
		
		if partnumber != '' and title != '':
			writing_dict[partnumber+'|'+title+'|'+brand_name+'|'+Land_url] = []
			spec_container = soup.select('.system_specs_container ul li')
			if len(spec_container) != 0:
				for specer in spec_container:
					attribute_name = specer.select('li .title')[0].text.strip()
					attribute_val = specer.select('li p')[0].text.strip() if len(specer.select('li p')) != 0 else ''
					writing_dict[partnumber+'|'+title+'|'+brand_name+'|'+Land_url].append((attribute_name, attribute_val))
		else:
			print('------ No Contents ------')
			writing_dict[brand_name+'|'+Land_url] = [('', 'No Content')]
	else:
		###################### Title ######################
		if len(soup.select('h2.singleModelTitle')) != 0:
			title = soup.select('h2.singleModelTitle')[0].text.strip()
		elif len(soup.select('h1.seo-title')) != 0:
			title = soup.select('h1.seo-title')[0].text.strip()
		elif len(soup.select('.headerTitle .titleSection')) != 0:
			title = soup.select('.headerTitle .titleSection')[0].text.strip()
		if '\n' in title:
			title = title.split('\n')[0]
		###################### Part Number ######################
		try:
			if len(soup.select('.partNumber')) != 0:
				try:
					partnumber = soup.select('.partNumber')[0].text.strip().split(':')[1].strip()
				except:
					partnumber = re.findall(r'.*?([0-9A-Z].*?)<', str(soup.select('.partNumber')[0]))
				if len(partnumber) != 0:
					partnumber = partnumber[0] if type(partnumber) == list else partnumber

			elif len(soup.select('.part-number')) != 0:
				partnumber = soup.select('.part-number')[0].text.strip().split(':')[1].strip()
			elif len(soup.select('.accessoriesDetail-partNumInfo span')) != 0:
				partnumber = soup.select('.accessoriesDetail-partNumInfo span')[0].text.strip().split(':')[1].strip()
			elif len(soup.select('.singleModelView button[data-productcode]')) != 0:
				partnumber = soup.select('.singleModelView button[data-productcode]')[0]['data-productcode']
		except:
			partnumber = soup.select('[name="productCode"]')[0]['value']
		###################### Brand ######################
		brand_name = soup.select('[name="brand"]')[0]['content'] if len(soup.select('[name="brand"]')) != 0 else ''

		###################### Specs ######################
		if len(soup.select('#product-details-variant-notavailable')) == 0:
			if partnumber != '' and title != '':
				writing_dict[partnumber+'|'+title+'|'+brand_name+'|'+Land_url] = []
				configure_content = soup.select('ul.configuratorItem-mtmTable li.configuratorItem-mtmTable-row')
				if len(configure_content) != 0:
					for configure in configure_content:
						attribute_name = configure.select('h4')[0]['data-term']
						attribute_val = configure.select('p.configuratorItem-mtmTable-description')[0].text.strip()
						writing_dict[partnumber+'|'+title+'|'+brand_name+'|'+Land_url].append((attribute_name, attribute_val))
				else:
					table_content = soup.select('table.techSpecs-table tbody tr')
					if len(table_content) != 0:
						for table_data in table_content:
							if len(table_data.select('td')) != 0:
								if len(table_data.select('td[colspan]')) == 0:
									attribute_name = table_data.select('td')[0].text.strip()
									attribute_val = table_data.select('td')[1].text.strip()
									writing_dict[partnumber+'|'+title+'|'+brand_name+'|'+Land_url].append((attribute_name, attribute_val))
			else:
				print('------ No Contents ------')
				writing_dict[brand_name+'|'+Land_url] = [('', 'No Content')]
		else:
			print('------ No info for the product ------')
			writing_dict[brand_name+'|'+Land_url] = [('', 'No productinfo')]


def land_page_execution(land_page_links, writing_dict):
	for prod_ind, Land_url in enumerate(land_page_links):
		prod_url = url_correction(Land_url)
		print(f'Processing [{prod_ind+1}/{len(land_page_links)}] : {prod_url}')
		
		techDriver.get(Land_url)
		time.sleep(2)
		land_soup = BeautifulSoup(techDriver.page_source, 'lxml')
		
		multi_content_page = land_soup.select('ol.tabbedBrowse-productListings li[data-code]')
		if len(multi_content_page) == 0:
			multi_content_page = land_soup.select('ul.skeleton_product li.product_item')
		#################################### Landing Port ####################################
		if len(multi_content_page) != 0:
			multi_product(land_soup, writing_dict, Land_url)
		else:
			single_product(land_soup, writing_dict, Land_url)

# category_product_urls = ['https://www.lenovo.com/fr/fr/p/laptops/thinkpad/thinkpadp/Thinkpad-P16-(16-inch-Intel)/21D6CTO1WWFR1']
# writing_dict = {'Laptops': {}}
# land_page_execution(category_product_urls, writing_dict['Laptops'])
# print(writing_dict)

def grid_list_page(grid_link, wrkstion_products):
	grid_page = requests.get(grid_link, headers=lenovo_headers)
	if grid_page.status_code == 200:
		grid_soup = BeautifulSoup(grid_page.text, 'lxml')
		grid_list = grid_soup.select('h3.seriesListings-title a')
		if len(grid_list) != 0:
			grid_products = [wrkstion_products.append(horizon['href']) for horizon in grid_list]

def common_list_page(list_link, product_urls):
	techDriver.get(list_link)
	time.sleep(3)

	if 'workstation' in list_link.lower():
		height = techDriver.execute_script("return document.body.scrollHeight;")
		scroll_by = 2000
		techDriver.execute_script("window.scrollBy(0, {});".format(scroll_by))
		time.sleep(5)
	common_list_soup = BeautifulSoup(techDriver.page_source, 'lxml')


	series_list_page = common_list_soup.select('ol.seriesListings')
	land_product_module = common_list_soup.select('.tabbedBrowse-module')
	land_product_module2 = common_list_soup.select('.banner_content_desc h2')
	list_product_count = common_list_soup.select('.dlp-filters__total-results')
	list_product_count2 = common_list_soup.select('.results .total')
	list_product_count3 = common_list_soup.select('#facetTop-count')

	selector_string = ''
	list_product_collector = ''
	if len(series_list_page) != 0:
		grid_list_page(list_link, product_urls)
	elif len(land_product_module) != 0:
		product_urls.append(list_link)
	elif len(land_product_module2) != 0:
		product_urls.append(list_link)
	elif len(list_product_count) != 0:
		selector_string = 'button.px-6.secondary-btn'
		list_product_collector = list_product_count
	elif len(list_product_count2) != 0:
		selector_string = 'button[data-tkey="loadMoreResults"]'
		list_product_collector = list_product_count2
	elif len(list_product_count3) != 0:
		selector_string = '.facetResults .facetResults-loadmore .loadmore-button'
		list_product_collector = list_product_count3

	if selector_string != '' and list_product_collector != '':
		list_product_collector = list_product_collector[0].text.strip()
		total_product_count = re.findall(r'(^\d*)',str(list_product_collector))[0]
		page_count = int(math.ceil(int(total_product_count)/20))
		
		for page_no in range(1, page_count):
			load_element = techDriver.find_elements(By.CSS_SELECTOR, selector_string)
			if len(load_element) != 0:
				try:
					load_element = WebDriverWait(techDriver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, selector_string)))
					try:
						load_element.click()
					except:
						techDriver.execute_script("arguments[0].click();", load_element)
					time.sleep(5)
				except:
					print(traceback.format_exc())
					pass
		height = techDriver.execute_script("return document.body.scrollHeight;")
		scroll_by = 150
		num_iterations = height//scroll_by
		for i in range(num_iterations):
			techDriver.execute_script("window.scrollBy(0, {});".format(scroll_by))
			time.sleep(0.1)

	list_items = techDriver.find_elements(By.CSS_SELECTOR, '.product__card__grid a.product__card__title-grid-link')
	list_items2 = techDriver.find_elements(By.CSS_SELECTOR, '.product_list li.product_item .product_title a')
	list_items3 = techDriver.find_elements(By.CSS_SELECTOR, '.facetResults .facetResults-item a.product_title')
	[product_urls.append(url_correction(item_links.get_attribute('href'))) for item_links in list_items if len(list_items) != 0]
	[product_urls.append(url_correction(item_links.get_attribute('href'))) for item_links in list_items2 if len(list_items2) != 0]
	[product_urls.append(url_correction(item_links.get_attribute('href'))) for item_links in list_items3 if len(list_items3) != 0]

def MajorExecutors(category_link, category_product_urls):
	global platform_domain
	techDriver.get(category_link)
	time.sleep(3)
	category_soup = BeautifulSoup(techDriver.page_source, 'lxml')
	category_response = Selector(text=techDriver.page_source)
	if platform_domain:
		if category_soup.select('.searchByType ul li a') if 'tablet' in category_link.lower() else category_soup.select('.searchBrandCards ul li a:nth-child(1)'):
			BrandListContent = category_soup.select('.searchByType ul li a') if 'tablet' in category_link.lower() else category_soup.select('.searchBrandCards ul li a:nth-child(1)') 
		# elif category_response.xpath('//h4[contains(text(),"유형")]/following-sibling::ul/li/a'):
		# 	BrandListContent = category_response.xpath('//h4[contains(text(),"유형")]/following-sibling::ul/li/a').getall()
		elif category_soup.select('h4[original_text="유형"]'):
			BrandListContent = category_soup.select('h4[original_text="유형"]')[0].find_next_sibling('ul').select('li a')
		else:
			BrandListContent = category_soup.select('.searchByType ul li a') if 'tablet' in category_link.lower() else category_soup.select('.searchBrandCards ul li a:nth-child(1)') 
	else:
		BrandListContent = category_soup.select('.container.byType a.cardByType') if 'tablet' in category_link.lower() else category_soup.select('.container.byBrands a.cardByBrands')

	if len(BrandListContent) != 0:
		for sub_type in BrandListContent:
			if len(sub_type['href']) != 0:
				sub_type_link = url_correction(sub_type['href'])
				if '%3A' in sub_type_link:
					print(f'Exceute Brand/Type : {re.findall(r"%3A(.*)",str(sub_type_link))[0]}')
				common_list_page(sub_type_link, category_product_urls)

def WorkStationExecutors(category_link, category_product_urls):
	techDriver.get(category_link)
	time.sleep(3)
	height = techDriver.execute_script("return document.body.scrollHeight;")
	scroll_by = 2000
	techDriver.execute_script("window.scrollBy(0, {});".format(scroll_by))
	time.sleep(5)
	category_soup = BeautifulSoup(techDriver.page_source, 'lxml')
	
	steroid_list = []
	series_content = category_soup.select('[class^=carousel-item-text] button a')
	underlist_product_content1 = category_soup.select('.product__card__grid a.product__card__title-grid-link')
	underlist_product_content2 = category_soup.select('.product_list li.product_item .product_title a')
	if len(series_content) != 0:
		[steroid_list.append(series_['href']) for series_ in series_content]
	print(f'Series products - {len(steroid_list)}')

	if len(underlist_product_content1) != 0 or len(underlist_product_content2) != 0:
		common_list_page(category_link, steroid_list)
	category_product_urls = list(dict.fromkeys(steroid_list))
	return category_product_urls


input_category = input('Enter the Category you need [ex. Laptops, Desktops,... & All ] : ')

error_text = ''
country_name = ''
for contry in lenovo_contries:
	if contry != None:
		contry_link = connect_domain+'/'+contry[0].lower()+'/'+contry[1].lower()+'/pc/'
		print(contry_link)
		#################### customized workbook module ############################
		customize_filename = f"Lenovo Category_{contry[0]}.xlsx"
		if os.path.exists(current_directory+customize_filename) == True:
			country_book = load_workbook(current_directory+customize_filename)
			lap_sheet = country_book['Laptop']
			desktop_sheet = country_book['Desktop']
			wrkstion_sheet = country_book['Workstations']
			tab_sheet = country_book['Tablets']
		else:
			country_book = openpyxl.Workbook()
			lap_sheet = country_book.active
			lap_sheet.title = 'Laptop'
			desktop_sheet = country_book.create_sheet('Desktop')
			wrkstion_sheet = country_book.create_sheet('Workstations')
			tab_sheet = country_book.create_sheet('Tablets')
			sheet_headers(lap_sheet)
			sheet_headers(desktop_sheet)
			sheet_headers(wrkstion_sheet)
			sheet_headers(tab_sheet)

		lap_sheet_row = lap_sheet.max_row+1
		desktop_sheet_row = desktop_sheet.max_row+1
		wrkstion_sheet_row = wrkstion_sheet.max_row+1
		tab_sheet_row = tab_sheet.max_row+1

		#################### customized workbook module ############################

		country_name = contry[0]
		home_page = requests.get(contry_link, headers=lenovo_headers)
		
		selective_category_links = []
		if home_page.status_code == 200:
			print('\n######################################################### [{}] #########################################################\n'.format(contry[0]))
			home_soup = BeautifulSoup(home_page.text, 'lxml')
			response = Selector(text=home_page.text)
			platform_domain = (home_soup.select('meta[name="platform"][content="Flash"]'))
			if platform_domain:
				print('# Platform : FLASH')
				pc_laptop= ''
				if response.xpath('//*[contains(@data-url,"laptops")]/a/@href|//*[contains(@data-name,"노트북")]/@href'):
					pc_laptop =[url for url in response.xpath('//*[contains(@data-url,"laptops")]/a/@href|//*[contains(@data-name,"노트북")]/@href').getall()if 'deal' not in url][0]
				if (pc_laptop) =='':
					pc_laptop = home_soup.select('.second_list li.second_list_item[data-url*="/laptops/"] a')
				if (pc_laptop) =='':
					pc_laptop = home_soup.select('.second_list li.second_list_item[data-url*="LAPTOP"] a')
				pc_desktops = home_soup.select('.second_list li.second_list_item[data-url*="desktops"] a')
				if len(pc_desktops) == 0:
					pc_desktops = home_soup.select('.second_list li.second_list_item[data-url*="DESKTOP"] a')
				pc_workstation = home_soup.select('.second_list li.second_list_item[data-url*="workstation"] a')
				if home_soup.select('.second_list li.second_list_item[data-url*="tablets"] a'):
					pc_tablets = home_soup.select('.second_list li.second_list_item[data-url*="tablets"] a')
				else:
					pc_tablets = response.xpath('//*[contains(@data-name,"태블릿")]/@href|//*[contains(@data-name,"平板電腦")]/@href').getall()
				if len(pc_tablets) == 0:
					pc_tablets = home_soup.select('.second_list li.second_list_item[data-url*="TABLET"] a')	
				pc_laptop = pc_laptop if len(pc_laptop) != 0 else 'no url'
				pc_desktops = pc_desktops[0]['href'] if len(pc_desktops) != 0 else 'no url'
				pc_workstation = pc_workstation[0]['href'] if len(pc_workstation) != 0 else 'no url'
				pc_tablets = pc_tablets[0]['href'] if len(pc_tablets) != 0 else 'no url'
			else:
				print('# Platform : HMC')
				dominant_root = home_soup.select('.o-mastheadModuleSuper__list li')
				if len(dominant_root) != 0:
					pc_laptop = dominant_root[0].select('.m-mastheadSubNav__list li a.m-subNav__link.icon-laptops')[0]['href']
					pc_desktops = dominant_root[0].select('.m-mastheadSubNav__list li a.m-subNav__link.icon-desktops')[0]['href']
					pc_workstation = dominant_root[0].select('.m-mastheadSubNav__list li a.m-subNav__link.icon-workstation')[0]['href']
					pc_tablets = dominant_root[0].select('.m-mastheadSubNav__list li a.m-subNav__link.icon-tablets')[0]['href']
				else:
					# if home_soup.select('.navBarSub a.icon-laptops'):
					# 	pc_laptop = home_soup.select('.navBarSub a.icon-laptops')[0]['href']
					# else:
     
					if response.xpath('//*[@class="icon-laptops"]/@href'):
						pc_laptop = response.xpath('//*[@class="icon-laptops"]/@href').get('')
					else:
						pc_laptop = response.xpath('//*[contains(@title,"Laptops")]/@href|//*[contains(@title,"Laptop")]/@href').get('')
					if response.xpath('//*[@class="icon-desktops"]/@href'):
						pc_desktops = response.xpath('//*[@class="icon-desktops"]/@href').get('')
					else:
						pc_desktops = response.xpath('//*[contains(@title,"Desktop")]/@href|//*[contains(@title,"PCs de escritorio")]/@href').get('')
					if response.xpath('//*[@class="icon-workstation"]/@href'):
						pc_workstation = response.xpath('//a[@class="icon-workstation"]/@href').get('')
					else:
						pc_workstation = response.xpath('//*[contains(@title,"Workstations")]/@href|//*[contains(@title,"Workstation")]/@href').get('')
					if response.xpath('//*[@class="icon-tablets"]/@href'):
						pc_tablets = response.xpath('//*[@class="icon-tablets"]/@href').get('')
					else:
						pc_tablets = response.xpath('//*[contains(@title,"Tablets")]/@href').get('')
					if response.xpath('//*[@class="icon-notebook"]/@href'):
						pc_tablets = response.xpath('//*[@class="icon-notebook"]/@href').get('')
					else:
						pc_tablets = response.xpath('//*[contains(@title,"Notebook")]/@href').get('')
					# pc_desktops = home_soup.select('.navBarSub a.icon-desktops')[0]['href']
					# pc_workstation = home_soup.select('.navBarSub a.icon-workstation')[0]['href']
					# pc_tablets = home_soup.select('.navBarSub a.icon-tablets')[0]['href']
			writing_dict = {}
			category_product_urls = []
			if 'laptop' in input_category.lower() or 'laptops' in input_category.lower():
				if 'laptop' in pc_laptop.lower() or 'notebook' in pc_laptop.lower() :
					writing_dict['Laptops'] = {}
					print('\nProcessing Category --------------------------- [Laptops] ---------------------------\n')
					MajorExecutors(url_correction(pc_laptop), category_product_urls)
					print('Total Laptops : ',len(category_product_urls))
					land_page_execution(category_product_urls, writing_dict['Laptops'])
			elif 'desk' in input_category.lower():
				if 'desktop' in pc_desktops.lower():
					writing_dict['Desktops'] = {}
					print('\nProcessing Category --------------------------- [Desktops] ---------------------------\n')
					MajorExecutors(url_correction(pc_desktops), category_product_urls)
					print('Total Desktops : ',len(category_product_urls))
					land_page_execution(category_product_urls, writing_dict['Desktops'])
			elif 'tablet' in input_category.lower():
				if 'tablet' in pc_tablets.lower():
					writing_dict['Tablets'] = {}
					print('\nProcessing Category --------------------------- [Tablets] ---------------------------\n')
					MajorExecutors(url_correction(pc_tablets), category_product_urls)
					print('Total Tablets : ',len(category_product_urls))
					land_page_execution(category_product_urls, writing_dict['Tablets'])
			elif 'work' in input_category.lower():
				if 'workstation' in pc_workstation.lower():
					writing_dict['Workstations'] = {}
					print('\nProcessing Category --------------------------- [Workstations] ---------------------------\n')
					category_product_urls = WorkStationExecutors(url_correction(pc_workstation), category_product_urls)
					print('Total Workstations : ',len(category_product_urls))
					land_page_execution(category_product_urls, writing_dict['Workstations'])
			elif 'all' in input_category.lower():
				lap_lists = []
				desk_lists = []
				tab_lists = []
				work_lists = []
				if 'laptop' in pc_laptop.lower() or 'notebook' in pc_laptop.lower():
					writing_dict['Laptops'] = {}
					print('\nProcessing Category --------------------------- [Laptops] ---------------------------\n')
					MajorExecutors(url_correction(pc_laptop), lap_lists)
					print('Total Laptops : ',len(lap_lists))
					land_page_execution(lap_lists, writing_dict['Laptops'])
				if 'desktop' in pc_desktops.lower():
					writing_dict['Desktops'] = {}
					print('\nProcessing Category --------------------------- [Desktops] ---------------------------\n')
					MajorExecutors(url_correction(pc_desktops), desk_lists)
					print('Total Desktops : ',len(desk_lists))
					land_page_execution(desk_lists, writing_dict['Desktops'])
				if 'tablet' in pc_tablets.lower():
					writing_dict['Tablets'] = {}
					print('\nProcessing Category --------------------------- [Tablets] ---------------------------\n')
					MajorExecutors(url_correction(pc_tablets), tab_lists)
					print('Total Tablets : ',len(tab_lists))
					land_page_execution(tab_lists, writing_dict['Tablets'])
				if 'workstation' in pc_workstation.lower():
					writing_dict['Workstations'] = {}
					print('\nProcessing Category --------------------------- [Workstations] ---------------------------\n')
					work_lists = WorkStationExecutors(url_correction(pc_workstation), work_lists)
					print('Total Workstations : ',len(work_lists))
					land_page_execution(work_lists, writing_dict['Workstations'])


			for category_key, category_val in writing_dict.items():
				for wkey, wval in category_val.items():
					part_number = wkey.split('|')[0] if len(wkey.split('|')) != 2 else ''
					part_title = wkey.split('|')[1] if len(wkey.split('|')) != 2 else ''
					part_brand = wkey.split('|')[2] if len(wkey.split('|')) != 2 else wkey.split('|')[0]
					part_url = wkey.split('|')[3] if len(wkey.split('|')) != 2 else wkey.split('|')[1]
					for write_val in wval:
						if category_key == 'Laptops':
							lap_sheet.cell(lap_sheet_row, 1).value = category_key
							lap_sheet.cell(lap_sheet_row, 2).value = contry[0]
							lap_sheet.cell(lap_sheet_row, 3).value = contry[1]
							lap_sheet.cell(lap_sheet_row, 4).value = part_brand
							lap_sheet.cell(lap_sheet_row, 5).value = part_title
							lap_sheet.cell(lap_sheet_row, 6).value = part_number
							lap_sheet.cell(lap_sheet_row, 7).value = write_val[0]
							lap_sheet.cell(lap_sheet_row, 8).value = write_val[1]
							lap_sheet.cell(lap_sheet_row, 9).value = part_url
							lap_sheet_row+=1
						if category_key == 'Desktops':
							desktop_sheet.cell(desktop_sheet_row, 1).value = category_key
							desktop_sheet.cell(desktop_sheet_row, 2).value = contry[0]
							desktop_sheet.cell(desktop_sheet_row, 3).value = contry[1]
							desktop_sheet.cell(desktop_sheet_row, 4).value = part_brand
							desktop_sheet.cell(desktop_sheet_row, 5).value = part_title
							desktop_sheet.cell(desktop_sheet_row, 6).value = part_number
							desktop_sheet.cell(desktop_sheet_row, 7).value = write_val[0]
							desktop_sheet.cell(desktop_sheet_row, 8).value = write_val[1]
							desktop_sheet.cell(desktop_sheet_row, 9).value = part_url
							desktop_sheet_row+=1
						if category_key == 'Tablets':
							tab_sheet.cell(tab_sheet_row, 1).value = category_key
							tab_sheet.cell(tab_sheet_row, 2).value = contry[0]
							tab_sheet.cell(tab_sheet_row, 3).value = contry[1]
							tab_sheet.cell(tab_sheet_row, 4).value = part_brand
							tab_sheet.cell(tab_sheet_row, 5).value = part_title
							tab_sheet.cell(tab_sheet_row, 6).value = part_number
							tab_sheet.cell(tab_sheet_row, 7).value = write_val[0]
							tab_sheet.cell(tab_sheet_row, 8).value = write_val[1]
							tab_sheet.cell(tab_sheet_row, 9).value = part_url
							tab_sheet_row+=1
						if category_key == 'Workstations':
							wrkstion_sheet.cell(wrkstion_sheet_row, 1).value = category_key
							wrkstion_sheet.cell(wrkstion_sheet_row, 2).value = contry[0]
							wrkstion_sheet.cell(wrkstion_sheet_row, 3).value = contry[1]
							wrkstion_sheet.cell(wrkstion_sheet_row, 4).value = part_brand
							wrkstion_sheet.cell(wrkstion_sheet_row, 5).value = part_title
							wrkstion_sheet.cell(wrkstion_sheet_row, 6).value = part_number
							wrkstion_sheet.cell(wrkstion_sheet_row, 7).value = write_val[0]
							wrkstion_sheet.cell(wrkstion_sheet_row, 8).value = write_val[1]
							wrkstion_sheet.cell(wrkstion_sheet_row, 9).value = part_url
							wrkstion_sheet_row+=1
				country_book.save(customize_filename)

time_off = time.strftime("%I:%M %p", time.localtime())
print('\n'+'End time: '+time_off)

elapsed_time = (time.time() - timer_start)
timeline = (time.strftime("Script time taken: %H:hr %M:min %S:sec ", time.gmtime(elapsed_time)))
timeliner = (time.strftime('%H hr: %M min: %S sec', time.gmtime(elapsed_time)))
print('\n'+timeline)