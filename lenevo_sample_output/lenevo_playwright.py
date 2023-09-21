import re
from os import getcwd, path
import pandas as pd
import openpyxl
from parsel import Selector
import asyncio
from playwright.async_api import async_playwright


async def main():
    """Xlsx Name WSD_Fresca_BI_February"""
    file_path = getcwd()
    file_name = input("Enter file name : ") + ".xlsx"
    output_filename = f"{file_path}\\{file_name}"
    wb = openpyxl.load_workbook(output_filename)
    # wb.active = wb['Vanity Grace']
    data_sheet = wb.active
    data_sheet["I1"] = "Checking_sku"
    data_sheet["J1"] = "Checking_url"
    data_sheet["K1"] = "Checking_product_availability"
    for i in range(1, data_sheet.max_row + 1):
        item = {}
        url = data_sheet.cell(row=i, column=1).value
        # url = data_sheet.cell(row=i, column=2).value
        # breakpoint()
        if url != None:
            async with async_playwright() as playwright:
                chromium = playwright.firefox
                browser = await chromium.launch(
                    args=["--start-maximized"], headless=False
                )
                page = await browser.new_page(
                    no_viewport=True, java_script_enabled=True
                )
                # try:
                    # """WSD_Fresca_BI_February"""
                page.set_default_timeout(0)
                response = await page.goto(url, wait_until="load")
                await page.wait_for_timeout(4000)
                await page.click('//button[@class="evidon-barrier-acceptbutton"]')
                response_text = Selector(text=await page.content())
                breakpoint()
                # response_text.xpath('//div[@class="lenovo_body"]/main[@class="main_content"]')
                response_text = Selector(text=await page.content())
                # with open('response_check.html','w') as f: f.write(await page.content())
                breakpoint()
                print(">>>>>>>>>>>>>>", page.url)

    #                 sku_check = bool(re.search(str(sku), await page.content()))
    #                 if response.status == 200 or sku_check == True:
    #                     sku_check = bool(re.search(str(sku), await page.content()))

    #                     data_sheet.cell(row=i, column=9).value = sku
    #                     data_sheet.cell(row=i, column=10).value = url

    #                     if sku_check == True:
    #                         print(">>>>>>>Status>>>>>>>", "Available")
    #                         product_availability = "Available"
    #                     else:
    #                         print(">>>>>>>Status>>>>>>>", " Not Available")
    #                         product_availability = "Not Available"
    #                     data_sheet.cell(row=i, column=11).value = product_availability
    #                 else:
    #                     print(">>>>>>>Status>>>>>>>", " Not Available")
    #                     product_availability = "Not Available"
    #                     data_sheet.cell(row=i, column=9).value = sku
    #                     data_sheet.cell(row=i, column=10).value = url
    #                     data_sheet.cell(row=i, column=11).value = product_availability
    #             except:
    #                 pass
    #             await page.close()
    #             await browser.close()
    #             await playwright.stop()
    #     else:
    #         print(">>>>>>>Status>>>>>>>", " Not Available")
    #         product_availability = "Not Available"
    #         data_sheet.cell(row=i, column=9).value = sku
    #         data_sheet.cell(row=i, column=10).value = url
    #         data_sheet.cell(row=i, column=11).value = product_availability
    # wb.save(output_filename)


asyncio.run(main())
